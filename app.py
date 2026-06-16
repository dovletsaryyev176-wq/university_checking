import io
from datetime import date
from flask import Flask, render_template, request, redirect, url_for, flash, send_file
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import db

app = Flask(__name__)
app.secret_key = 'university-secret-key'

# Fixed-width field slices for the student file format
FIELDS = [
    ('student_number', slice(0,   10)),
    ('first_name',     slice(10,  22)),
    ('last_name',      slice(22,  34)),
    ('book_type',      slice(34,  35)),
    ('direction1',     slice(35,  39)),
    ('direction2',     slice(39,  43)),
    ('direction3',     slice(43,  47)),
    ('direction4',     slice(47,  51)),
    ('turkmen',        slice(51,  71)),
    ('english',        slice(71,  101)),
    ('informatics',    slice(101, 116)),
    ('history',        slice(116, 131)),
    ('jemgyyet',       slice(131, 146)),
    ('economics',      slice(146, 161)),
    ('biology',        slice(161, 176)),
    ('chemistry',      slice(176, 196)),
    ('mathematics',    slice(196, 216)),
    ('physics',        slice(216, 236)),
    ('zehin',          slice(236, 251)),
]

# Offsets of each subject within the flat answer key string
SUBJECT_KEY_SLICES = [
    ('turkmen',     slice(0,   20)),
    ('english',     slice(20,  50)),
    ('informatics', slice(50,  65)),
    ('history',     slice(65,  80)),
    ('jemgyyet',    slice(80,  95)),
    ('economics',   slice(95,  110)),
    ('biology',     slice(110, 125)),
    ('chemistry',   slice(125, 145)),
    ('mathematics', slice(145, 165)),
    ('physics',     slice(165, 185)),
    ('zehin',       slice(185, 200)),
]

SUBJECT_LABELS = {
    'turkmen':     'Türkmen dili',
    'english':     'Iňlis dili',
    'informatics': 'Informatika',
    'history':     'Taryh',
    'jemgyyet':    'Jemgyýet',
    'economics':   'Ykdysadyýet',
    'biology':     'Biologiýa',
    'chemistry':   'Himiýa',
    'mathematics': 'Matematika',
    'physics':     'Fizika',
    'zehin':       'Zehin',
}

def parse_line(line):
    return {name: line[s].strip() for name, s in FIELDS}

def compare_student(student, key_str):
    result = {}
    total_correct = total_wrong = total_no_answer = 0
    for subject, key_slice in SUBJECT_KEY_SLICES:
        student_answers = (student[subject] or '').strip()
        key_answers = key_str[key_slice]
        chars = []
        correct = wrong = no_answer = 0
        for s_char, k_char in zip(student_answers, key_answers):
            if s_char == k_char:
                chars.append((s_char, 'correct'))
                correct += 1
            elif s_char == '#':
                chars.append((s_char, 'no_answer'))
                no_answer += 1
            else:
                chars.append((s_char, 'wrong'))
                wrong += 1
        for _ in range(len(student_answers), len(key_answers)):
            chars.append((' ', 'wrong'))
            wrong += 1
        result[subject] = {
            'chars': chars, 'correct': correct,
            'wrong': wrong, 'no_answer': no_answer,
            'total': len(key_answers),
        }
        total_correct += correct
        total_wrong   += wrong
        total_no_answer += no_answer
    score = total_correct - total_wrong / 4
    result['_total'] = {
        'correct':   total_correct,
        'wrong':     total_wrong,
        'no_answer': total_no_answer,
        'score':     round(score, 2),
    }
    return result

@app.route('/')
def index():
    students = db.get_all_students()
    answer_keys = db.get_all_answer_keys()
    compare = request.args.get('compare') == '1'
    comparison = {}
    if compare:
        for s in students:
            key_row = answer_keys.get(s['book_type'])
            if key_row:
                cmp = compare_student(s, key_row['answers'])
                comparison[s['id']] = cmp
                db.upsert_result(s['id'], cmp)
    return render_template('index.html', students=students, answer_keys=answer_keys,
                           compare=compare, comparison=comparison)

@app.route('/upload', methods=['POST'])
def upload():
    file = request.files.get('file')
    if not file or file.filename == '':
        flash('Faýl saýlanmady.', 'warning')
        return redirect(url_for('index'))

    raw = file.read()
    try:
        content = raw.decode('utf-8')
    except UnicodeDecodeError:
        content = raw.decode('windows-1251')

    added = 0
    skipped = []

    for line in content.splitlines():
        if len(line) < 251:
            continue
        student = parse_line(line)
        if not student['student_number']:
            continue
        if db.student_exists(student['student_number']):
            skipped.append(student['student_number'])
        else:
            db.insert_student(student)
            added += 1

    if added:
        flash(f'Goşulan talyplaryň sany: {added}', 'success')
    if skipped:
        flash(f'Eýýäm bar (geçildi): {", ".join(skipped)}', 'warning')
    if not added and not skipped:
        flash('Faýlda laýyk setirler tapylmady.', 'danger')

    return redirect(url_for('index'))

@app.route('/upload_key', methods=['POST'])
def upload_key():
    book_type = request.form.get('book_type', '').strip().upper()
    if book_type not in ('A', 'B'):
        flash('Kitabyň görnüşi nädogry. Diňe A ýa-da B bolmalydyr.', 'danger')
        return redirect(url_for('index'))

    file = request.files.get('file')
    if not file or file.filename == '':
        flash('Faýl saýlanmady.', 'warning')
        return redirect(url_for('index'))

    raw = file.read()
    try:
        content = raw.decode('utf-8')
    except UnicodeDecodeError:
        content = raw.decode('windows-1251')

    line = content.splitlines()[0].strip() if content.strip() else ''
    if not line or not all(c in 'ABCD' for c in line):
        flash('Faýl A, B, C, D simwollaryndan ybarat setir saklamalydyr.', 'danger')
        return redirect(url_for('index'))

    db.upsert_answer_key(book_type, line)
    db.delete_results_for_book_type(book_type)
    flash(f'{book_type} kitaby üçin jogap açary saklandy. Köne netijeler öçürildi.', 'success')
    return redirect(url_for('index'))

@app.route('/delete/<int:student_id>', methods=['POST'])
def delete(student_id):
    db.delete_student(student_id)
    flash('Talyp öçürildi.', 'success')
    return redirect(url_for('index'))

@app.route('/export')
def export():
    students    = db.get_all_students()
    answer_keys = db.get_all_answer_keys()
    subjects    = [s for s, _ in SUBJECT_KEY_SLICES]

    wb = Workbook()
    ws = wb.active
    ws.title = 'Netijeler'

    green  = PatternFill('solid', fgColor='C6EFCE')
    red    = PatternFill('solid', fgColor='FFC7CE')
    gray   = PatternFill('solid', fgColor='D9D9D9')
    yellow = PatternFill('solid', fgColor='FFEB9C')
    bold   = Font(bold=True)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)

    INFO_COLS = ['№', 'Belgisi', 'Ady', 'Familiýasy', 'Kitap',
                 'Ugur 1', 'Ugur 2', 'Ugur 3', 'Ugur 4']
    n_info = len(INFO_COLS)

    # Row 1: merged group headers
    ws.append(INFO_COLS + [SUBJECT_LABELS[s] for s in subjects for _ in range(3)]
              + ['Jemi', None, None, 'Bal'])

    # Row 2: sub-headers В / Н / #
    ws.append([''] * n_info
              + ['D', 'Ý', '#'] * len(subjects)
              + ['D', 'Ý', '#', ''])

    # Merge info columns across rows 1-2
    for col in range(1, n_info + 1):
        ws.merge_cells(start_row=1, start_column=col, end_row=2, end_column=col)

    # Merge each subject group (3 cols) in row 1
    for i, _ in enumerate(subjects):
        c = n_info + 1 + i * 3
        ws.merge_cells(start_row=1, start_column=c, end_row=1, end_column=c + 2)

    # Merge "Итого" group and merge "Балл" across rows 1-2
    total_start = n_info + 1 + len(subjects) * 3
    ws.merge_cells(start_row=1, start_column=total_start, end_row=1, end_column=total_start + 2)
    score_col = total_start + 3
    ws.merge_cells(start_row=1, start_column=score_col, end_row=2, end_column=score_col)

    # Style header rows
    for r in range(1, 3):
        for cell in ws[r]:
            cell.font      = bold
            cell.alignment = center

    # Color sub-header cells (row 2)
    for i in range(len(subjects)):
        c = n_info + 1 + i * 3
        ws.cell(2, c).fill     = green
        ws.cell(2, c + 1).fill = red
        ws.cell(2, c + 2).fill = gray

    ws.cell(2, total_start).fill     = green
    ws.cell(2, total_start + 1).fill = red
    ws.cell(2, total_start + 2).fill = gray
    ws.cell(1, score_col).fill = yellow
    ws.cell(2, score_col).fill = yellow

    # Data rows
    for i, s in enumerate(students, 1):
        key_row = answer_keys.get(s['book_type'])
        cmp = compare_student(s, key_row['answers']) if key_row else None

        row = [i, s['student_number'], s['first_name'], s['last_name'],
               s['book_type'], s['direction1'], s['direction2'],
               s['direction3'], s['direction4']]

        if cmp:
            for subj in subjects:
                row += [cmp[subj]['correct'], cmp[subj]['wrong'], cmp[subj]['no_answer']]
            t = cmp['_total']
            row += [t['correct'], t['wrong'], t['no_answer'], t['score']]
            db.upsert_result(s['id'], cmp)
        else:
            row += [''] * (len(subjects) * 3 + 4)

        ws.append(row)

        dr = i + 2
        for j in range(len(subjects)):
            c = n_info + 1 + j * 3
            ws.cell(dr, c).fill     = green
            ws.cell(dr, c + 1).fill = red
            ws.cell(dr, c + 2).fill = gray
        ws.cell(dr, total_start).fill     = green
        ws.cell(dr, total_start + 1).fill = red
        ws.cell(dr, total_start + 2).fill = gray
        ws.cell(dr, score_col).fill = yellow

    # Column widths
    for col in range(1, n_info + 1):
        ws.column_dimensions[get_column_letter(col)].width = 11
    for col in range(n_info + 1, score_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = 5
    ws.column_dimensions[get_column_letter(score_col)].width = 8

    ws.freeze_panes = 'A3'

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f'results_{date.today()}.xlsx'
    return send_file(output, as_attachment=True, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

if __name__ == '__main__':
    db.init()
    app.run(host='0.0.0.0', port=8000, debug=True)
